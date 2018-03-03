import openpyxl
import pickle
import time

start_time = time.time()

wb = openpyxl.Workbook()
ws = wb.active

with open("dumps/movie_corate_matrix.pickle", "rb") as movie_corate_matrix_dump:
    movie_corate_matrix = pickle.load(movie_corate_matrix_dump)

# ws.cell(column = 1,row = 1, value = "movie id vs movie id")
for i in range(1682):
    for j in range(i + 1, 1682):
        print i, j
        ws.cell(row=i + 1, column=j + 1, value=movie_corate_matrix[i][j])
wb.save(filename="demo.xls")
print "Time:", time.time() - start_time
